import os

from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

import openpyxl
from openpyxl.styles import Font, Alignment, colors

import glob
import datetime


#Data to be written on excel
encabezado = {
    1 :'id_licitacion',
    2 :'nro_contrato',
    3 :'fecha_firma',
    4 :'nombre_empresa',
    5 :'nombre_licitacion',
    6 :'monto_adjudicado',
    7 :'monto_ampliacion',
    8 :'monto_fonacide',
    9 :'periodo',
    10:'plurianual',
    11:'categoria',
    12:'rescindido',
    13:'link_contrato'
    }


#Data to upload to the cloud (Hernandarias district)
contratos_id = '14Ab05yUVz0-VH8vU83G0cIVfIgsgvUSa'
years_id = {2014: '11k3wyoWUn19kKsF4HjTvY1klpBFoDNTH',
            2019: '1GPJwZsigvVV1rMuKWohFrXN8sgzDykIP',
            2018: '1nOiDDtSTXpvXfVFIIWjosnfEzLBhGExh',
            2017: '1aq4CglVACVgU67pAY1q31F9NWFLQzxxp',
            2016: '1uynLY219y3Zj9BY1Q3pYLJWGBfjcdXgx',
            2015: '1EQHzifuBGjH4hIkDwApiOm7vAGaB1OPb'}

#Setting style of cells in excel
font = Font(name = 'Arial',
            sz = 10)

alignment = Alignment(horizontal = 'center',
                      vertical = 'center',
                      text_rotation = 0,
                      wrap_text = False,
                      shrink_to_fit = False,
                      indent = 0)
def overwrite(sheet, contratos):
    '''This function will overwrite the rows that
    are already in the excel file
    '''
    len(contratos)
    #We copy because we need to give the contracts that are not in the excel file
    copy_contratos = []
    
    
    
    for contrato in contratos:
        column = [str(sheet.cell(i, 2).value) for i in range(1, 742)]
        copy_contratos.append(contrato)
        #IF it's on the table, lets look through them
        condition = True
        if not(contrato['nro_contrato'] in column):
            condition = False
        else:
            row = column.index(contrato['nro_contrato'])+1

        is_in_excel = False
        
        while condition:
            
            if int(sheet.cell(row, 1).value) != contrato['id_licitacion']:
                column[column.index(contrato['nro_contrato'])] = "lol"
            else:
                is_in_excel = True
                break
            if not (contrato['nro_contrato'] in column):
                break
            else:
                row = column.index(contrato['nro_contrato'])+1
        #print("Control:")
        #print(is_in_excel)
        
        if is_in_excel:
            
            for i in list(encabezado.keys()):
                
                sheet.cell(row, i).value =  contrato[encabezado[i]]
                sheet.cell(row, i).font = font
                sheet.cell(row, i).alignment = alignment


            copy_contratos.remove(contrato)
    #print(len(copy_contratos))
    return copy_contratos

def writehere(sheet, contratos, n):
    if len(contratos) == 0:
        return None
    for contrato in contratos:
        for i in list(encabezado.keys()):
            sheet.cell(n, i).value = contrato[encabezado[i]]
            sheet.cell(n, i).font = font
            sheet.cell(n, i).alignment = alignment

        n+=1
    
    
        
def find_min(sheet, m):
    '''This function will try to find the earliest place in the sheet to write the data'''
    if m == 0:
        #if nothing to add
        return None
    n = 0
    condition = False
    while not condition:
        n += 1
        
        condition = all([not bool(sheet.cell(i,1).value) for i in range(n, m+n)])

        
    return n


def checkcredentials():
    #Check credentials. If an error is raised, run gencredentials.py
    try:
        gauth = GoogleAuth()
        gauth.LoadCredentialsFile('mycreds.txt')
        gauth.Authorize()
        
    except Exception as e:
        
        import credentialsgen
        
    finally:
        drive = GoogleDrive(gauth)
        
    return drive

def upload_to_cloud(drive, year_id):

    
    
    #List all folders inside the year
    drive_folders = drive.ListFile({'q':"'{}' in parents and trashed=false".format(year_id)}).GetList()

    #List all remote folders
    name_folders = [fold['title'] for fold in drive_folders]
    #Splitting the characteristics of the folders
    name_folders_spplited = [fold.split(" ") for fold in name_folders]
    name_folders_id = [fold['id'] for fold in drive_folders]

    #List all local folders
    local_folders = os.listdir()
    #Fixing Windows protocol
    local_folders = [fold.replace("-", "/") for fold in local_folders]

    cwd = os.getcwd()

    #Try to match local and
    for folder in local_folders:
        
        #This variable will check whether the folder was found on the could or not
        found = False
        #Getting the characteristics of both folders, to compare them
        char_local = folder.split(" ")

        #Try to find corresponding folder in drive
        for chars in name_folders_spplited:
            if chars[0] == char_local[0] and chars[1] == char_local[1]:
                
                #Store matching id of the drive folder
                match_id = name_folders_id[name_folders_spplited.index(chars)]
                '''
                Debugging
                print(name_folders_id)
                print(name_folders_spplited)
                print(match_id)
                print(chars)
                print(char_local)
                '''
                #Delete already charged values for optimization (could be discarded
                name_folders_spplited.remove(chars)
                name_folders_id.remove(match_id)
                found = True
                break
        
        #If not found, create a folder
        if not found:
            new_folder = drive.CreateFile({'title':folder,
                                           'parents':[{'id':year_id}],
                                           'mimeType': 'application/vnd.google-apps.folder'})
            new_folder.Upload()
            
            match_id = new_folder['id']
            
        #Enter to the folder and update the files if they are not in drive
        os.chdir("./"+folder.replace("/", "-"))

        #List year folder's elements on Drive
        year_folder = drive.ListFile({'q':"'{}' in parents and trashed=false".format(match_id)}).GetList()
        
        elements = [element['title'].lower() for element in year_folder]

        #----------------------------------------------------------
        #----------------------------------------------------------
        #----------------------------------------------------------
        #Is codigo de contratacion in the folder?
        is_cc = False
        for  title in elements:
            if "codigo" in title:
                is_cc = True
        #If not, upload it
        if not is_cc:
            file = drive.CreateFile({'title':'Codigo de Contratacion',
                                    'parents':[{'id':match_id}],
                                    'mimeType':'application/pdf'})

            file.SetContentFile("Codigo de Contratacion.pdf")
            file.Upload()
            print("Uploaded Codigo de Contratacion")
        #----------------------------------------------------------
        #----------------------------------------------------------
        #----------------------------------------------------------
        
        #Is the contract in the folder?
        is_con = False
        for  title in elements:
            if "contrato" in title:
                is_con = True

        #If not, upload it
        if not is_con:
            
            contract = os.listdir()
            contract = [pdf for pdf in contract if "contrato" in pdf.lower()][0]

            file = drive.CreateFile({'title':contract.split(".")[0],
                                    'parents':[{'id':match_id}],
                                    'mimeType':'application/pdf'})
            file.SetContentFile(contract)
            file.Upload()
            print("Uploaded {}".format(contract))
        #----------------------------------------------------------
        #----------------------------------------------------------
        #----------------------------------------------------------
        #Does the contract have ampliacion?
        
        if os.path.isfile("./Ampliacion (CC).pdf"):
           

            #Is the ampliacion cc in the folder?

            is_amp = False

            for title in elements:
                if "ampliacion (cc)" in title:
                    is_amp = True
                    break

            #If not, upload it
            if not is_amp:

                ampliacion_pdf = os.listdir()
                ampliacion_pdf = [pdf for pdf in ampliacion_pdf if "ampliacion" in pdf.lower()][0]
    
                file = drive.CreateFile({'title':ampliacion_pdf.split(".")[0],
                                    'parents':[{'id':match_id}],
                                    'mimeType':'application/pdf'})
                file.SetContentFile(ampliacion_pdf)

                file.Upload()
                print("Uploaded {}".format(ampliacion_pdf))
        #----------------------------------------------------------
        #----------------------------------------------------------
        #----------------------------------------------------------
        #Does the Contract have Pliego de Bases y condiciones?

        if any(["bases y condiciones" in title.lower() for title in os.listdir()]):

            #Is the pbc in the remote folder?

            is_pbc = False
            
            for title in elements:
                if "Bases" in title or "bases" in title or "pbc" in title:
                    is_pbc = True
                    break

            #If not, upload it

            if not is_pbc:

                pbc_file = os.listdir()
                pbc_file = [file for file in pbc_file
                            if ("bases" in file.lower() or "pbc" in file.lower())][0]

                mimeType = pbc_file.split(".")[-1]

                file = drive.CreateFile({'title':pbc_file.split(".")[0],
                                         'parents':[{'id':match_id}],
                                         'mimeType':'application/{}'.format(mimeType)})
                file.SetContentFile(pbc_file)

                file.Upload()
                print("Uploaded {}".format(pbc_file))

        #Reset Current Working directory
        os.chdir(cwd)
        
def get_links(drive, contratos, year_id):

    year_folder = drive.ListFile({'q':"'{}' in parents and trashed=false".format(year_id)}).GetList()

    for drive_contract in year_folder:

        for local_contract in contratos:
            if str(local_contract['nro_contrato']) == drive_contract['title'].split(" ")[0] and \
               str(local_contract['id_licitacion']) == drive_contract['title'].split(" ")[1]:

                match_id = drive_contract['id']

                files_in_drive = drive.ListFile({'q':"'{}' in parents and trashed=false".format(match_id)}).GetList()

                contract_link = [file for file in files_in_drive if (("contrato" in file['title'].lower()) and \
                                                                     not("ampliacion" in file['title'].lower()))][0]
                contract_link = 'https://drive.google.com/open?id={}'.format(contract_link['id']) 

                #Beware!
                local_contract.update({'link_contrato':contract_link})


    return contratos
                
        
    
    

def main(contratos = [], year = datetime.datetime.now().year, exceptions = [], municipio = "Hernandarias", excel = True,
         upto= True):

    drive = checkcredentials()
    
    #Lets upload the files
    os.chdir("./Temps{}/{}".format(municipio, year))

  
    year_id = years_id[year]

    #Now Upload The Append Files, if neccesary
    
    if upto:
        upload_to_cloud(drive, year_id)

    contratos = get_links(drive, contratos, year_id)
    contratos_complete = contratos
    #Step back
    os.chdir(os.path.dirname(os.getcwd()))

    if excel:
        #Exclude Exceptions
        copy_contratos = []
        for contrato in contratos:
            if not ("{} {}".format(contrato['id_licitacion'], contrato['nro_contrato']) in exceptions):
                copy_contratos.append(contrato)

        contratos = copy_contratos
        del copy_contratos
            
        
        hern_id = '1B3-A1aznOtuSijnf-OCvZDXPKBxYEgVr'
        

        #Check if there is an excel file in path destination, if so, remove it.
        if len(glob.glob(os.getcwd() + "\\*.xlsx")) > 0:
            for file in glob.glob(os.getcwd() + "\\*.xlsx"):
                os.remove(file)

        #Download the file
        list2 = drive.ListFile(
            {'q':"'%s' in parents and trashed=false" % hern_id}).GetList()
        xlsxData = [a for a in list2 if "Resumen Contratos" in a['title']][1]

        file = drive.CreateFile({'id':xlsxData['id']})
        

  
        file.GetContentFile(xlsxData['title'])

        #Open the file on both reading and writing mode
        book = openpyxl.load_workbook(xlsxData['title'])
        sheet = book['SoloContratos']

      
        #Update the data, if necessary
        remaining_contratos = overwrite(sheet, contratos)
      
        #See how much space do we need to store the data:
        n = find_min(sheet, len(remaining_contratos))
     
        #Write the remaining contratos
        writehere(sheet, remaining_contratos, n)

        
      
        #Update the file
        book.save(xlsxData['title'])

        #Upload to Drive
        file.SetContentFile(xlsxData['title'])



        
        file.Upload()
    
    os.chdir(os.path.dirname(os.getcwd()))
    
  
    return contratos_complete
    
    
if __name__ == '__main__':

    pass
